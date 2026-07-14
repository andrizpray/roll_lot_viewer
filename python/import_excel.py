"""Import roll lots or paper sheets from an Excel file into the database."""
import os
import sys
import datetime

sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from db import execute_query, execute_many, get_connection
from config import IMPORT_BATCH_SIZE, UPLOAD_DIR


# Column mapping: Excel header → DB column
# Support both formats: with spaces and without
COLUMN_MAP = {
    # Common columns (Roll & Sheet)
    "Lot ID": "lot_id",
    "LotID": "lot_id",
    "Item ID": "item_id",
    "ItemID": "item_id",
    "Weight": "weight",
    "Description": "description_raw",
    "Date": "source_tr_date",
    "TrDate": "source_tr_date",
    "Time": "source_tr_time",
    "TrTime": "source_tr_time",
    "Comments": "comments",
    
    # Roll-specific columns
    "endqty": "weight",
    "Paper Type": "papertype",
    "PaperType": "papertype",
    "Gramature": "gramature",
    "Play Bond": "playbond",
    "PlayBond": "playbond",
    "Width": "width",
    "Rew ID": "rew_id",
    "RewID": "rew_id",
    "Grade": "grade",
    "Diameter": "diameter",
    "Thickness": "thickness",
    
    # Sheet-specific columns
    "Qty": "qty",
    "Keterangan": "keterangan",
    "Qty_Pack": "content_pack",
    "Dimension": "dimension",
}

# Columns that exist in roll_lot_histories table
HISTORY_COLUMNS = [
    "lot_id", "item_id", "weight", "papertype", "gramature",
    "playbond", "width", "rew_id", "grade", "comments",
    "diameter", "thickness", "description_raw",
    "source_tr_date", "source_tr_time", "import_batch_id",
]


def parse_description(description):
    """
    Parse description to extract papertype, gramature, playbond, and width.
    FOR ROLL LOTS ONLY.
    
    Example: "B Kraft BK125 E150 690"
    - papertype: "B Kraft" (first 1-2 words)
    - gramature: "BK125" (word before playbond)
    - playbond: "E150" (word before width)
    - width: "690" (last word)
    
    Returns: dict with keys papertype, gramature, playbond, width (values can be None)
    """
    import re
    
    result = {
        'papertype': None,
        'gramature': None,
        'playbond': None,
        'width': None,
    }
    
    if not description or not isinstance(description, str):
        return result
    
    # Clean up: remove text in parentheses (e.g., "(Item Blocked)")
    # This prevents suffixes from interfering with parsing
    description = re.sub(r'\s*\([^)]*\)', '', description)
    
    # Split by whitespace and filter empty strings
    parts = [p.strip() for p in description.split() if p.strip()]
    
    if len(parts) == 0:
        return result
    
    # Width is the last token (if it looks like a number)
    if len(parts) >= 1:
        last = parts[-1]
        # Check if it's numeric or contains digits
        if any(c.isdigit() for c in last):
            result['width'] = last
            parts = parts[:-1]  # Remove width from parts
    
    # PlayBond is the second-to-last token (pattern: E + number)
    if len(parts) >= 1:
        playbond_candidate = parts[-1]
        result['playbond'] = playbond_candidate
        parts = parts[:-1]
    
    # Gramature is the third-to-last token (pattern: BK/NK + number)
    if len(parts) >= 1:
        gramature_candidate = parts[-1]
        result['gramature'] = gramature_candidate
        parts = parts[:-1]
    
    # PaperType is everything remaining (1-2 words at the start)
    if len(parts) >= 1:
        result['papertype'] = ' '.join(parts)
    
    return result


def parse_description_sheet(description):
    """
    Parse description to extract papertype, gramature, and dimension.
    FOR PAPER SHEETS ONLY.
    
    Example: "BK350 590x840"
    - First word "BK350": 
      - Prefix "BK" = papertype code → "B Kraft"
      - Number "350" = gramature
    - Second word "590x840" = dimension
    
    Edge case: "40007433 CB350G 770x650"
    - Skip numeric prefix (40007433)
    - Parse "CB350G" as gramature (with G suffix)
    
    Paper Type Codes:
    - BK = B Kraft
    - GB = Gray Board
    - BSP = B Kraft Sheet PE
    - CB = ChipBoard
    - CO = CoreBoard
    - YB = YellowBoard
    
    Returns: dict with keys papertype, gramature, dimension
    """
    import re
    
    result = {
        'papertype': None,
        'gramature': None,
        'dimension': None,
    }
    
    if not description or not isinstance(description, str):
        return result
    
    # Paper type code mapping
    papertype_map = {
        'BK': 'B Kraft',
        'GB': 'Gray Board',
        'BSP': 'B Kraft Sheet PE',
        'CB': 'ChipBoard',
        'CO': 'CoreBoard',
        'YB': 'YellowBoard',
        'BB': 'Brown Board',
        'DS': 'Duplex Sheet',
        'KBD': 'Kraft Back Duplex',
        'NK': 'Natural Kraft',
        'NSP': 'Natural Kraft Sheet PE',
        'DPS': 'Duplex PE Sheet',
    }
    
    parts = description.strip().split()
    
    if len(parts) == 0:
        return result
    
    # Skip first word if it's pure numeric (edge case: item code prefix)
    start_idx = 0
    if parts[0].isdigit():
        start_idx = 1
    
    if start_idx >= len(parts):
        return result
    
    # First meaningful word: extract papertype code + gramature
    # Pattern: 2-3 letters + numbers + optional letter suffix (e.g., BK350, CB350G, GB1000)
    first_word = parts[start_idx]
    match = re.match(r'^([A-Z]{2,3})(\d+[A-Z]?)', first_word, re.IGNORECASE)
    
    if match:
        code = match.group(1).upper()
        gramature_full = match.group(2)
        
        # Map code to full paper type name
        if code in papertype_map:
            result['papertype'] = papertype_map[code]
        else:
            result['papertype'] = code  # Use code as-is if not in map
        
        # Store full gramature with code (e.g., "BK350", "CB350G")
        result['gramature'] = code + gramature_full
    
    # Next word: dimension (pattern: NUMxNUM or NUM X NUM)
    next_idx = start_idx + 1
    if next_idx < len(parts):
        dimension_candidate = parts[next_idx]
        # Check if it matches dimension pattern (digits x digits, case insensitive)
        if re.search(r'\d+\s*[xX×]\s*\d+', dimension_candidate):
            result['dimension'] = dimension_candidate
    
    return result


def import_roll_lots(job_id, filepath):
    """Import roll lots or paper sheets from an Excel file."""
    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"Import file not found: {filepath}")

    # Determine target table and file path from import_jobs
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute("SELECT type FROM import_jobs WHERE id = %s", (job_id,))
        result = cur.fetchone()
        job_type = result[0] if result else "roll"
        stored_path = result[1] if result else None
    finally:
        conn.close()

    # Resolve actual file path
    if stored_path and os.path.isfile(stored_path):
        filepath = stored_path
    elif stored_path:
        # storage_path is relative hash — join with UPLOAD_DIR
        candidate = os.path.join(UPLOAD_DIR, stored_path)
        if os.path.isfile(candidate):
            filepath = candidate

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Workbook has no active sheet")

    # Read headers from first row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Auto-detect type from headers if PHP got it wrong
    detected = detect_type_from_headers(headers)
    if detected != job_type:
        print(f"[import] job {job_type} -> detected '{detected}' from headers, overriding")
        job_type = detected
        # Update DB
        conn2 = get_connection()
        try:
            conn2.execute("UPDATE import_jobs SET type = ? WHERE id = ?", (job_type, job_id))
            conn2.commit()
        finally:
            conn2.close()

    table = "paper_sheets" if job_type == "sheet" else "roll_lots"
    column_map = SHEET_COLUMN_MAP if job_type == "sheet" else ROLL_COLUMN_MAP

    # Handle sheet mode with positional indexing (duplicate headers)
    if job_type == "sheet":
        return _import_sheet_rows(job_id, filepath, wb, ws, headers)

    # Roll mode: map headers to DB columns
    db_columns = []
    header_to_col = {}
    for h in headers:
        if h in COLUMN_MAP:
            db_columns.append(COLUMN_MAP[h])
            header_to_col[h] = COLUMN_MAP[h]

    if not db_columns:
        raise ValueError(f"No matching columns found. Excel headers: {headers}")

    # Build INSERT SQL — use INSERT ... ON CONFLICT (upsert) for PostgreSQL
    placeholders = ", ".join(["%s"] * len(db_columns))
    columns = ", ".join([f'"{c}"' for c in db_columns])
    update_cols = ", ".join([f'"{c}" = EXCLUDED."{c}"' for c in db_columns])
    insert_sql = (
        f'INSERT INTO {table} ({columns}, import_batch_id) '
        f'VALUES ({placeholders}, %s) '
        f'ON CONFLICT (lot_id) DO UPDATE SET {update_cols}, import_batch_id = EXCLUDED.import_batch_id'
    )

    # Process rows
    total = 0
    success = 0
    failed = 0
    errors = []  # (row_number, lot_id, description_raw, reason)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue

        total += 1
        row_number = total + 1  # +1 for header row

        try:
            # Map row values to DB columns
            data = dict(zip(headers, row))
            values = []
            for h in headers:
                if h in COLUMN_MAP:
                    val = data.get(h)
                    # Normalize datetime objects to strings for SQLite
                    if isinstance(val, datetime.time):
                        val = val.strftime("%H:%M:%S")
                    elif isinstance(val, datetime.date):
                        val = val.strftime("%Y-%m-%d")
                    elif isinstance(val, datetime.datetime):
                        val = val.strftime("%Y-%m-%d %H:%M:%S")
                    # Convert empty strings to None for cleaner data
                    elif val == "":
                        val = None
                    values.append(val)

            # Extract lot_id and description for processing
            lot_id = data.get("Lot ID") or data.get("LotID")
            description_raw = data.get("Description")
            
            # Parse description to fill missing fields (different logic for Roll vs Sheet)
            if description_raw:
                # Create a dict mapping db_column to index in values array
                col_index = {COLUMN_MAP[h]: i for i, h in enumerate(headers) if h in COLUMN_MAP}
                
                if table == "paper_sheets":
                    # Sheet parsing: Description → papertype, gramature, dimension
                    parsed = parse_description_sheet(description_raw)
                    
                    if "papertype" in col_index and not values[col_index["papertype"]]:
                        values[col_index["papertype"]] = parsed['papertype']
                    
                    if "gramature" in col_index and not values[col_index["gramature"]]:
                        values[col_index["gramature"]] = parsed['gramature']
                    
                    if "dimension" in col_index and not values[col_index["dimension"]]:
                        values[col_index["dimension"]] = parsed['dimension']
                
                else:
                    # Roll parsing: Description → papertype, gramature, playbond, width
                    parsed = parse_description(description_raw)
                    
                    if "papertype" in col_index and not values[col_index["papertype"]]:
                        values[col_index["papertype"]] = parsed['papertype']
                    
                    if "gramature" in col_index and not values[col_index["gramature"]]:
                        values[col_index["gramature"]] = parsed['gramature']
                    
                    if "playbond" in col_index and not values[col_index["playbond"]]:
                        values[col_index["playbond"]] = parsed['playbond']
                    
                    if "width" in col_index and not values[col_index["width"]]:
                        values[col_index["width"]] = parsed['width']

            # --- Snapshot history (only for roll_lots, not sheets) ---
            if table == "roll_lots" and lot_id:
                _snapshot_existing(lot_id, job_id)

            # Append job_id as import_batch_id
            values.append(job_id)

            # Insert single row
            conn = get_connection()
            try:
                cur = conn.cursor()
                cur.execute(insert_sql, values)
                conn.commit()
            finally:
                conn.close()

            success += 1

        except Exception as exc:
            failed += 1
            lot_id_val = data.get("Lot ID") if 'data' in dir() else None
            desc_val = data.get("Description") if 'data' in dir() else None
            errors.append((row_number, lot_id_val, desc_val, str(exc)))

        # Update progress every IMPORT_BATCH_SIZE rows
        if total % IMPORT_BATCH_SIZE == 0:
            _update_progress(job_id, total, success, failed)

    wb.close()

    # Log errors to import_errors table
    if errors:
        _log_errors(job_id, errors)

    # Final update
    _update_progress(job_id, total, success, failed, completed=True)

    print(f"[import] job {job_id}: imported {success}/{total} rows "
          f"({failed} errors) from {filepath}")
    return success


def _snapshot_existing(lot_id, job_id):
    """Copy existing roll_lot record to roll_lot_histories before overwrite."""
    conn = get_connection()
    try:
        cur = conn.cursor()
        # Check if this lot_id already exists
        cur.execute("SELECT * FROM roll_lots WHERE lot_id = %s", (lot_id,))
        row = cur.fetchone()
        if row is None:
            return  # No existing record to snapshot

        # Get column names from cursor description
        col_names = [desc[0] for desc in cur.description]
        existing = dict(zip(col_names, row))

        # Build history insert with only columns that exist in history table
        now = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        history_values = []
        history_cols = []

        for col in HISTORY_COLUMNS:
            if col in existing:
                history_cols.append(col)
                history_values.append(existing[col])

        # Add archived_at and created_at/updated_at
        history_cols.extend(["archived_at", "created_at", "updated_at"])
        history_values.extend([now, now, now])

        placeholders = ", ".join(["%s"] * len(history_cols))
        col_str = ", ".join([f'"{c}"' for c in history_cols])

        cur.execute(
            f"INSERT INTO roll_lot_histories ({col_str}) VALUES ({placeholders})",
            history_values,
        )
        conn.commit()
    finally:
        conn.close()


def _log_errors(job_id, errors):
    """Write per-row errors to the import_errors table."""
    now = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    insert_sql = (
        "INSERT INTO import_errors "
        "(import_batch_id, row_number, lot_id, description_raw, reason, created_at, updated_at) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s)"
    )
    data = [
        (job_id, row_num, lot_id, desc_raw, reason, now, now)
        for row_num, lot_id, desc_raw, reason in errors
    ]
    try:
        execute_many(insert_sql, data)
    except Exception as exc:
        # Don't let error-logging failures crash the import
        print(f"[import] WARNING: failed to log {len(errors)} errors: {exc}")


def _update_progress(job_id, total, success, failed, completed=False):
    """Update job progress in import_jobs table."""
    conn = get_connection()
    try:
        cur = conn.cursor()
        if completed:
            cur.execute(
                "UPDATE import_jobs SET total_rows = %s, success_count = %s, "
                "failed_count = %s, status = 'completed' WHERE id = %s",
                (total, success, failed, job_id),
            )
        else:
            cur.execute(
                "UPDATE import_jobs SET total_rows = %s, success_count = %s, "
                "failed_count = %s WHERE id = %s",
                (total, success, failed, job_id),
            )
        conn.commit()
    finally:
        conn.close()
