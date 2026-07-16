"""Export roll lots or paper sheets from the database into an Excel file."""
import json
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from db import execute_query
from config import EXPORT_DIR, MAX_EXPORT_ROWS

# Column configs per mode
ROLL_HEADERS = [
    "No.", "Lot ID", "Item ID", "Weight", "Paper Type", "Gramature",
    "Play Bond", "Width", "Rew ID", "Grade", "Comments",
    "Diameter", "Thickness", "Description", "Date", "Time",
]
ROLL_COLUMNS = [
    None, "lot_id", "item_id", "weight", "papertype", "gramature",
    "playbond", "width", "rew_id", "grade", "comments",
    "diameter", "thickness", "description_raw", "source_tr_date", "source_tr_time",
]

SHEET_HEADERS = [
    "No.", "Lot ID", "Item ID", "Weight", "Paper Type", "Gramature",
    "Dimension", "Content Pack", "Description", "Date", "Time",
]
SHEET_COLUMNS = [
    None, "lot_id", "item_id", "weight", "papertype", "gramature",
    "dimension", "content_pack", "description_raw", "source_tr_date", "source_tr_time",
]

# Numeric columns that should be stored as numbers, empty if no value
NUMERIC_COLUMNS = {"width", "grade"}
GRADE3_FILL = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")  # yellow highlight


def export_roll_lots(job_id, filters=None, mode="roll"):
    """Export roll lots or paper sheets to an Excel file."""
    filters = filters or {}
    table = "paper_sheets" if mode == "sheet" else "roll_lots"
    headers = SHEET_HEADERS if mode == "sheet" else ROLL_HEADERS
    db_columns = SHEET_COLUMNS if mode == "sheet" else ROLL_COLUMNS

    # Build query dynamically from filters
    where_clauses = []
    params = []

    if filters.get("item_id"):
        where_clauses.append("item_id = %s")
        params.append(filters["item_id"])

    if filters.get("grade"):
        # Support comma-separated multi-grade (e.g. "1,2,3")
        grades = filters["grade"]
        if isinstance(grades, str):
            grades = [g.strip() for g in grades.split(",") if g.strip()]
        elif not isinstance(grades, list):
            grades = [str(grades)]
        if len(grades) == 1:
            where_clauses.append("grade = %s")
            params.append(grades[0])
        elif grades:
            placeholders = ", ".join(["%s"] * len(grades))
            where_clauses.append(f"grade IN ({placeholders})")
            params.extend(grades)

    if filters.get("papertype"):
        where_clauses.append("papertype LIKE %s")
        params.append(f"%{filters['papertype']}%")

    if filters.get("gramature"):
        where_clauses.append("gramature LIKE %s")
        params.append(f"%{filters['gramature']}%")

    if filters.get("width"):
        where_clauses.append("width LIKE %s")
        params.append(f"%{filters['width']}%")

    if mode == "sheet" and filters.get("dimension"):
        where_clauses.append("dimension LIKE %s")
        params.append(f"%{filters['dimension']}%")

    if filters.get("lot_id"):
        where_clauses.append("lot_id LIKE %s")
        params.append(f"%{filters['lot_id']}%")

    if filters.get("date_from"):
        where_clauses.append("source_tr_date >= %s")
        params.append(filters["date_from"])

    if filters.get("date_to"):
        where_clauses.append("source_tr_date <= %s")
        params.append(filters["date_to"])

    # Batch mode: specific lot_ids
    if filters.get("lot_ids"):
        lot_ids = filters["lot_ids"]
        if lot_ids:
            placeholders = ", ".join(["%s"] * len(lot_ids))
            where_clauses.append(f"lot_id IN ({placeholders})")
            params.extend(lot_ids)

    sql = f"SELECT * FROM {table}"
    if where_clauses:
        sql += " WHERE " + " AND ".join(where_clauses)
    sql += f" ORDER BY lot_id LIMIT {MAX_EXPORT_ROWS}"

    rows = execute_query(sql, params if params else None)

    if not rows:
        raise RuntimeError("No data found for export")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Roll Lots" if mode == "roll" else "Paper Sheets"

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(db_columns, 1):
            if col_name is None:
                # No. column = row number
                ws.cell(row=row_idx, column=col_idx, value=row_idx - 1)
                continue
            value = row.get(col_name)
            if col_name in NUMERIC_COLUMNS:
                # Store as number, empty if no value
                value = int(value) if value is not None and value != "" else None
            ws.cell(row=row_idx, column=col_idx, value=value)

        # Highlight entire row if grade == 3
        if row.get("grade") is not None:
            try:
                if int(row["grade"]) == 3:
                    for c in range(1, len(db_columns) + 1):
                        ws.cell(row=row_idx, column=c).fill = GRADE3_FILL
            except (ValueError, TypeError):
                pass

    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    # Save file
    os.makedirs(EXPORT_DIR, exist_ok=True)
    filename = f"export_job_{job_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    wb.close()

    print(f"[export] job {job_id}: wrote {len(rows)} rows ({mode}) to {filepath}")
    return filepath
